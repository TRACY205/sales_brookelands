# myapp/views.py
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Sum
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from .models import Expense, Sale
from decimal import Decimal
from django.utils import timezone
import json

# --------------------------
# Price list for sales
# --------------------------
PRICE_LIST = {
    "Water": {
        "1L (R)": 10,
        "1L + B": 30,
        "1L (R)": 10,
        "5L (R)": 70,
        "5L Bottle": 150,
        "10L (R)": 130,
        "10L Bottle": 250,
        "20L (R)": 250,
        "20L  Bottle": 500,
        "20L (Hard) + water": 1500,
    },
    "Gas": {
        "Insta Gas 6kg": 1000,
        "Insta Gas 13kg": 3500,
        "Pro Gas 6kg": 1000,
        "Pro Gas 13kg": 3500,
        "Wajiko 6kg": 1000,
        "Wajiko 13kg": 3500,
    }
}

# --------------------------
# Landing page
# --------------------------
def landing(request):
    return render(request, "landing.html")

# --------------------------
# User registration
# --------------------------
def register_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password1 = request.POST.get("password1")
        password2 = request.POST.get("password2")

        if password1 != password2:
            messages.error(request, "Passwords do not match.")
            return redirect("register")

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect("register")

        User.objects.create_user(username=username, password=password1)
        messages.success(request, "Account created successfully! You can now login.")
        return redirect("login")

    return render(request, "register.html")

# --------------------------
# User login
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib import messages

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            if user.is_staff:  # Admin goes to admin dashboard
                return redirect("admin_dashboard")
            else:  # Normal user goes to user dashboard
                return redirect("user_dashboard")
        else:
            messages.error(request, "Invalid username or password")
            return redirect("login")

    return render(request, "login.html")

# --------------------------
# Logout
# --------------------------
def logout_view(request):
    logout(request)
    return redirect("login")

# --------------------------
# User dashboard
# --------------------------
from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
import json
from .models import Sale
@login_required
def user_dashboard(request):
    if request.method == "POST":
        category = request.POST.get("category")
        item = request.POST.get("item") or request.POST.get("local_item")
        quantity = request.POST.get("quantity")
        payment_method = request.POST.get("payment_method")
        payment_status = request.POST.get("payment_status")
        delivery_place = request.POST.get("delivery_place")  # ✅ capture delivery

        if not category or not item or not quantity or not payment_method or not payment_status:
            messages.error(request, "⚠️ Please fill all required fields.")
            return redirect("user_dashboard")

        # Automatically get price from PRICE_LIST
        item_price = Decimal(PRICE_LIST.get(item, 0))
        total_price = item_price * Decimal(quantity)

        Sale.objects.create(
            user=request.user,
            category=category,
            item=item,
            quantity=int(quantity),
            price=total_price,
            payment_method=payment_method,
            payment_status=payment_status,
            delivery_place=delivery_place if payment_status == "Delivery" else "",  # ✅ save only if Delivery
        )
        messages.success(request, f"✅ Order submitted successfully! {item} - {quantity} @ {item_price} = {total_price}")
        return redirect("user_dashboard")

    sales = Sale.objects.filter(user=request.user).order_by("-id")

    return render(request, "user_dashboard.html", {
        "sales": sales,
        "PRICE_LIST_JSON": json.dumps(PRICE_LIST),
    })

# --------------------------
# Admin dashboard
@login_required
def admin_dashboard(request):
    # Prevent unauthorized access
    if not request.user.is_superuser:
        messages.error(request, "🚫 Unauthorized access.")
        return redirect("user_dashboard")

    # Admin expenses (latest first by ID)
    expenses = Expense.objects.all().order_by("-id")
    total_expenses = expenses.aggregate(total=Sum("amount_paid"))["total"] or 0

    # User sales/orders (latest first by ID)
    sales = Sale.objects.all().order_by("-id")
    total_sales = sales.aggregate(total=Sum("price"))["total"] or 0

    context = {
        "expenses": expenses,
        "total_expenses": total_expenses,
        "sales": sales,
        "total_sales": total_sales,
    }
    return render(request, "admin_dashboard.html", context)








# --------------------------
# Add sale (user)
# --------------------------
from decimal import Decimal

@login_required
def add_sale(request):
    if request.method == "POST":
        category = request.POST.get("category")
        item = request.POST.get("item")
        quantity = request.POST.get("quantity")
        payment_method = request.POST.get("payment_method")

        try:
            quantity = int(quantity)
        except:
            quantity = 0

        unit_price = PRICE_LIST.get(item)
        if category and item and quantity > 0 and unit_price is not None and payment_method:
            total_price = Decimal(unit_price) * quantity

            # ✅ Store clean decimal value in DB
            Sale.objects.create(
                user=request.user,
                category=category,
                item=item,
                quantity=quantity,
                price=total_price,
                payment_method=payment_method
            )
            return redirect("user_dashboard")

    return render(request, "add_sale.html", {"PRICE_LIST": PRICE_LIST})


# --------------------------
# Add expense (admin)
# --------------------------
@login_required
def add_expense(request):
    if request.method == "POST":
        date = request.POST.get("date")
        paid_to = request.POST.get("paid_to")
        charged_to = request.POST.get("charged_to")
        description = request.POST.get("description")
        receipt_no = request.POST.get("receipt_no")
        sponsor = request.POST.get("sponsor")
        amount_injected = request.POST.get("amount_injected")
        amount_paid = request.POST.get("amount_paid")
        bank_charges = request.POST.get("bank_charges")
        running_balance = request.POST.get("running_balance")

        if not date or not paid_to or not charged_to or not description or not amount_injected or not amount_paid:
            messages.error(request, "Please fill all required fields.")
            return redirect("add_expense")

        try:
            amount_injected = float(amount_injected)
            amount_paid = float(amount_paid)
            bank_charges = float(bank_charges) if bank_charges else 0
            running_balance = float(running_balance) if running_balance else 0
        except ValueError:
            messages.error(request, "Amounts must be numbers.")
            return redirect("add_expense")

        Expense.objects.create(
            user=request.user,
            date=date,
            paid_to=paid_to,
            charged_to=charged_to,
            description=description,
            receipt_no=receipt_no,
            sponsor=sponsor,
            amount_injected=amount_injected,
            amount_paid=amount_paid,
            bank_charges=bank_charges,
            running_balance=running_balance
        )

        messages.success(request, "Expense added successfully!")
        return redirect("admin_dashboard")

    return render(request, "add_expense.html")

# --------------------------
# Upload expenses (Excel/CSV)
# --------------------------
@login_required
def upload_expense(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]
        # handle Excel/CSV upload logic here
        messages.success(request, "Expenses uploaded successfully!")
        return redirect("admin_dashboard")
    return render(request, "upload_expense.html")

# --------------------------
# Admin PDF reports
# --------------------------
# myapp/views.py






# myapp/views.py
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from .models import Expense, Sale

# ------------------- Excel Expenses -------------------
def admin_expenses_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="expenses_report.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses Report"

    headers = [
        "Date", "Paid To", "Charged To", "Description",
        "Receipt No", "Sponsor", "Amount Injected",
        "Amount Paid", "Bank Charges", "Running Balance"
    ]
    ws.append(headers)

    expenses = Expense.objects.all().order_by("date")

    running_balance = 0
    total_injected = 0
    total_paid = 0
    total_charges = 0

    for exp in expenses:
        injected = getattr(exp, 'amount_injected', 0) or 0
        paid = getattr(exp, 'amount_paid', 0) or 0
        charges = getattr(exp, 'bank_charges', 0) or 0

        running_balance += injected - paid - charges
        total_injected += injected
        total_paid += paid
        total_charges += charges
        ws.append([
    exp.date.strftime("%d/%m/%y") if exp.date else "",  # <-- comma added here
    getattr(exp, 'paid_to', ''),
    getattr(exp, 'charged_to', ''),
    getattr(exp, 'description', ''),
    getattr(exp, 'receipt_no', ''),
    getattr(exp, 'sponsor', ''),
    float(injected),
    float(paid),
    float(charges),
    float(running_balance),
])


        
  
        

    ws.append([
        "", "", "", "TOTALS", "", "",
        float(total_injected),
        float(total_paid),
        float(total_charges),
        float(running_balance),
    ])

    wb.save(response)
    return response

# ------------------- Excel Sales with Totals by Payment Method -------------------
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Font, numbers
from django.utils.dateparse import parse_date
from .models import Sale


# =========================
# 📊 Sales Report (HTML page)
# =========================
def sales_report(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    sales = Sale.objects.all().order_by("date")

    # Filter by date range
    if start_date:
        sales = sales.filter(date__gte=parse_date(start_date))
    if end_date:
        sales = sales.filter(date__lte=parse_date(end_date))

    total_amount = sales.aggregate(Sum("price"))["price__sum"] or 0

    return render(request, "admin_sales_report.html", {
        "sales": sales,
        "total_amount": total_amount,
    })
# =========================
# 📥 Export Sales to Excel (Safe Version)
# =========================
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, numbers
from django.utils.dateparse import parse_date
from datetime import datetime
from .models import Sale

def admin_sales_excel(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    sales = Sale.objects.all().order_by("date")

    # ✅ Filter by dates
    if start_date:
        sales = sales.filter(date__gte=parse_date(start_date))
    if end_date:
        sales = sales.filter(date__lte=parse_date(end_date))

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Headers
    headers = ["Date", "Item", "Quantity", "Price", "Payment Method"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Totals
    total_price = 0
    total_cash = 0
    total_mpesa = 0

    total_delivery = 0
    delivery_qty = 0

    total_r_amount = 0
    total_r_liters = 0  # already there

    total_gas_amount = 0
    gas_qty = 0

    total_bottle_amount = 0
    bottle_qty = 0

    if not sales.exists():
        ws.append(["No sales data available"])
    else:
        for s in sales:
            # ✅ Safe date handling
            sale_date = ""
            if s.date:
                if isinstance(s.date, str):
                    try:
                        sale_date = datetime.strptime(s.date, "%Y-%m-%d").strftime("%d/%m/%y")
                    except ValueError:
                        sale_date = s.date
                else:
                    sale_date = s.date.strftime("%d/%m/%y")

            ws.append([
                sale_date,
                s.item,
                s.quantity,
                float(s.price),
                f"{s.payment_method} ({s.payment_status})",
            ])
            ws.cell(row=ws.max_row, column=4).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

            # Totals tracking
            price_val = float(s.price or 0)
            qty_val = int(s.quantity or 0)
            total_price += price_val

            # Liter tracking for (R)
            if "(R)" in str(s.item):
                total_r_liters += qty_val
                total_r_amount += price_val

            # Payment method totals
            if s.payment_method and s.payment_method.lower() == "cash":
                total_cash += price_val
            elif s.payment_method and s.payment_method.lower() == "mpesa":
                total_mpesa += price_val

            # Delivery total
            if s.payment_status and s.payment_status.lower() == "delivery":
                total_delivery += price_val
                delivery_qty += qty_val

            # Gas total
            if "gas" in str(s.item).lower() or "wajiko" in str(s.item).lower():
                total_gas_amount += price_val
                gas_qty += qty_val

            # Bottle total
            if "bottle" in str(s.item).lower():
                total_bottle_amount += price_val
                bottle_qty += qty_val

        # Totals section
        ws.append([])
        ws.append(["TOTAL (R only)", "", total_r_liters, total_r_amount, ""])
        ws.append(["TOTAL Delivery", "", delivery_qty, total_delivery, ""])
        ws.append(["TOTAL Gas", "", gas_qty, total_gas_amount, ""])
        ws.append(["TOTAL Bottle", "", bottle_qty, total_bottle_amount, ""])
        ws.append(["", "Cash Total", "", total_cash, ""])
        ws.append(["", "MPesa Total", "", total_mpesa, ""])
        ws.append(["", "Overall Total", "", total_price, ""])

        # Bold totals
        for row in ws.iter_rows(min_row=ws.max_row-6, max_row=ws.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.font = Font(bold=True)

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="sales_report.xlsx"'
    wb.save(response)
    return response
# myapp/views.py



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Sale
from .forms import SaleForm


@login_required
def edit_order(request, pk):
    sale = get_object_or_404(Sale, pk=pk)

    if request.method == "POST":
        form = SaleForm(request.POST, instance=sale)
        if form.is_valid():
            form.save()
            messages.success(request, "Order updated successfully!")
            return redirect("dashboard")
    else:
        form = SaleForm(instance=sale)

    return render(request, "edit_order.html", {"form": form, "sale": sale})



# myapp/views.py
from django.shortcuts import render, get_object_or_404, redirect
from .models import Expense
from .forms import ExpenseForm

def expenses_list(request):
    expenses = Expense.objects.all()
    return render(request, "expenses_list.html", {"expenses": expenses})

def edit_expense(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == "POST":
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            return redirect("admin_dashboard")

    else:
        form = ExpenseForm(instance=expense)
    return render(request, "edit_expense.html", {"form": form, "expense": expense})  




from django.shortcuts import render, get_object_or_404, redirect
from .models import Sale
from .forms import SaleForm  # make sure you have this form

@login_required
def edit_sale(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)
    if request.method == "POST":
        form = SaleForm(request.POST, instance=sale)
        if form.is_valid():
            form.save()
            return redirect('admin_dashboard')  # back to dashboard
    else:
        form = SaleForm(instance=sale)
    return render(request, "edit_sale.html", {"form": form, "sale": sale})



from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import redirect, get_object_or_404
from .models import Sale, Expense


@login_required
def delete_orders(request):
    if request.method == "POST":
        order_ids = request.POST.getlist("selected_orders")
        if order_ids:
            Sale.objects.filter(id__in=order_ids).delete()
            messages.success(request, f"{len(order_ids)} order(s) deleted successfully.")
        else:
            messages.error(request, "No orders were selected for deletion.")
    return redirect("dashboard")   # user dashboard


@login_required
def delete_sales(request):
    if request.method == "POST":
        sale_ids = request.POST.getlist("selected_sales")
        if sale_ids:
            Sale.objects.filter(id__in=sale_ids).delete()
            messages.success(request, f"{len(sale_ids)} sale(s) deleted successfully.")
        else:
            messages.error(request, "No sales were selected for deletion.")
    return redirect("admin_dashboard")   # admin dashboard


@login_required
def delete_expenses(request):
    if request.method == "POST":
        expense_ids = request.POST.getlist("selected_expenses")
        if expense_ids:
            Expense.objects.filter(id__in=expense_ids).delete()
            messages.success(request, f"{len(expense_ids)} expense(s) deleted successfully.")
        else:
            messages.error(request, "No expenses were selected for deletion.")
    return redirect("admin_dashboard")   # ✅ go back to admin dashboard    here?
